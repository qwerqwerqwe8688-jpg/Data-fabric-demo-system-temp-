# backend/collectors/file_collector.py
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any
import logging
from datetime import datetime
import re
import hashlib
import json
import sqlite3
import openpyxl

from backend.models.metadata import DataAsset, Column, DataRow, Sheet, Database
from .base_collector import BaseMetadataCollector


class FileCollector(BaseMetadataCollector):
    def __init__(self, base_path: str, sample_rows: int = 100):
        self.base_path = Path(base_path)
        self.sample_rows = sample_rows  # 采样行数，避免数据过大

    def test_connection(self) -> bool:
        return self.base_path.exists()

    def _make_safe_id(self, raw: str) -> str:
        """将任意路径/字段名映射为 URL-safe 字符串"""
        return re.sub(r'[^0-9A-Za-z._-]', '_', str(raw))

    def _calculate_row_hash(self, row_data: Dict[str, Any]) -> str:
        """计算行数据的哈希值"""
        row_str = json.dumps(row_data, sort_keys=True, ensure_ascii=False)
        return hashlib.md5(row_str.encode('utf-8')).hexdigest()

    def collect_metadata(self) -> List[DataAsset]:
        assets: List[DataAsset] = []
        now = datetime.now().isoformat()

        if not self.base_path.exists():
            logging.warning(f"Base path does not exist: {self.base_path}")
            return assets

        # 收集CSV文件
        for pattern in ["*.csv", "*.txt"]:
            for file_path in self.base_path.rglob(pattern):
                assets.extend(self._collect_csv_metadata(file_path, now))

        # 收集Excel文件
        for pattern in ["*.xlsx", "*.xls"]:
            for file_path in self.base_path.rglob(pattern):
                assets.extend(self._collect_excel_metadata(file_path, now))

        # 收集SQLite数据库文件 - 增强模式匹配
        sqlite_patterns = ["*.db", "*.sqlite", "*.sqlite3", "*.db3"]
        for pattern in sqlite_patterns:
            for file_path in self.base_path.rglob(pattern):
                logging.info(f"发现SQLite数据库文件: {file_path}")
                sqlite_assets = self._collect_sqlite_metadata(file_path, now)
                assets.extend(sqlite_assets)
                logging.info(f"从 {file_path.name} 采集到 {len(sqlite_assets)} 个资产")

        # 立即写入图数据库
        try:
            from backend.services.graph_service import GraphService
            gs = GraphService("bolt://localhost:7687", "neo4j", "password")
            for ast in assets:
                gs.create_asset(ast)
            gs.close()
        except Exception as e:
            logging.warning(f"写入图数据库失败: {e}")

        # 统计各类资产数量
        asset_types = {}
        for asset in assets:
            asset_types[asset.type] = asset_types.get(asset.type, 0) + 1

        logging.info(f"✅ FileCollector 已写入 {len(assets)} 个资产")
        for asset_type, count in asset_types.items():
            logging.info(f"  - {asset_type}: {count}个")

        return assets

    def _collect_csv_metadata(self, file_path: Path, timestamp: str) -> List[DataAsset]:
        """收集CSV文件元数据"""
        assets = []
        columns = self.get_csv_columns(file_path)
        if not columns:
            return assets

        # 1. 文件资产
        safe_file_id = self._make_safe_id(file_path.stem)
        file_asset = DataAsset(
            id=f"file.{safe_file_id}",
            name=file_path.name,
            type="file",
            description=f"CSV数据文件: {file_path}",
            owner="文件采集器",
            tags=["csv", "数据文件"],
            created_time=timestamp,
            updated_time=timestamp
        )
        assets.append(file_asset)

        # 2. 列资产
        for col_name in columns:
            safe_col_id = self._make_safe_id(col_name)
            col_asset = Column(
                id=f"file.{safe_file_id}.{safe_col_id}",
                name=col_name,
                type="column",
                data_type="string",
                description=f"列: {col_name} in {file_path.name}",
                owner="文件采集器",
                tags=["column", "数据列"],
                created_time=timestamp,
                updated_time=timestamp
            )
            assets.append(col_asset)

        # 3. 行级资产收集
        row_assets = self._collect_csv_row_metadata(file_path, safe_file_id, columns, timestamp)
        assets.extend(row_assets)

        return assets

    def _collect_excel_metadata(self, file_path: Path, timestamp: str) -> List[DataAsset]:
        """收集Excel文件元数据"""
        assets = []

        try:
            # 1. 文件资产
            safe_file_id = self._make_safe_id(file_path.stem)
            file_asset = DataAsset(
                id=f"excel.{safe_file_id}",
                name=file_path.name,
                type="file",
                description=f"Excel数据文件: {file_path}",
                owner="文件采集器",
                tags=["excel", "数据文件"],
                created_time=timestamp,
                updated_time=timestamp
            )
            assets.append(file_asset)

            # 2. 读取Excel文件
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # 获取列名（第一行）
                columns = []
                for col_idx in range(1, sheet.max_column + 1):
                    cell_value = sheet.cell(row=1, column=col_idx).value
                    if cell_value:
                        columns.append(str(cell_value))
                    else:
                        columns.append(f"Column_{col_idx}")

                if not columns:
                    continue

                # 3. 工作表资产
                safe_sheet_id = self._make_safe_id(sheet_name)
                sheet_asset = Sheet(
                    id=f"excel.{safe_file_id}.sheet.{safe_sheet_id}",
                    name=sheet_name,
                    type="sheet",
                    description=f"Excel工作表: {sheet_name}",
                    owner="文件采集器",
                    tags=["excel", "sheet", "工作表"],
                    created_time=timestamp,
                    updated_time=timestamp,
                    file_id=f"excel.{safe_file_id}",
                    sheet_name=sheet_name,
                    row_count=sheet.max_row - 1,  # 减去标题行
                    column_count=sheet.max_column
                )
                assets.append(sheet_asset)

                # 4. 列资产
                for col_name in columns:
                    safe_col_id = self._make_safe_id(col_name)
                    col_asset = Column(
                        id=f"excel.{safe_file_id}.sheet.{safe_sheet_id}.{safe_col_id}",
                        name=col_name,
                        type="column",
                        data_type="string",
                        description=f"列: {col_name} in {sheet_name}",
                        owner="文件采集器",
                        tags=["column", "数据列"],
                        created_time=timestamp,
                        updated_time=timestamp
                    )
                    assets.append(col_asset)

                # 5. 行级资产收集
                row_assets = self._collect_excel_row_metadata(sheet, safe_file_id, safe_sheet_id, columns, timestamp)
                assets.extend(row_assets)

            workbook.close()

        except Exception as e:
            logging.warning(f"Excel文件处理失败 {file_path}: {e}")

        return assets

    def _collect_sqlite_metadata(self, file_path: Path, timestamp: str) -> List[DataAsset]:
        """收集SQLite数据库元数据"""
        assets = []

        try:
            # 验证SQLite文件有效性
            if not self._is_valid_sqlite_file(file_path):
                logging.warning(f"无效的SQLite文件: {file_path}")
                return assets

            # 1. 数据库资产
            safe_db_id = self._make_safe_id(file_path.stem)
            db_asset = Database(
                id=f"sqlite.{safe_db_id}",
                name=file_path.name,
                type="database",
                description=f"SQLite数据库: {file_path}",
                owner="文件采集器",
                tags=["sqlite", "数据库"],
                created_time=timestamp,
                updated_time=timestamp,
                file_path=str(file_path),
                table_count=0,
                connection_string=f"sqlite:///{file_path}"
            )
            assets.append(db_asset)

            # 2. 连接数据库获取表信息
            conn = sqlite3.connect(file_path)
            cursor = conn.cursor()

            # 获取所有表
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()
            db_asset.table_count = len(tables)

            logging.info(f"在数据库 {file_path.name} 中发现 {len(tables)} 个表: {[table[0] for table in tables]}")

            for table_info in tables:
                table_name = table_info[0]
                safe_table_id = self._make_safe_id(table_name)

                # 3. 表资产
                table_asset = DataAsset(
                    id=f"sqlite.{safe_db_id}.table.{safe_table_id}",
                    name=table_name,
                    type="table",
                    description=f"SQLite表: {table_name}",
                    owner="文件采集器",
                    tags=["sqlite", "table", "数据表"],
                    created_time=timestamp,
                    updated_time=timestamp
                )
                assets.append(table_asset)

                # 4. 获取列信息和数据类型
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns_info = cursor.fetchall()

                # 获取表的前几行数据以推断数据类型
                cursor.execute(f"SELECT * FROM {table_name} LIMIT 5")
                sample_data = cursor.fetchall()
                column_names = [col[1] for col in columns_info]

                # 推断数据类型
                inferred_types = self._infer_column_types(sample_data, column_names)

                # 5. 列资产
                for col_info in columns_info:
                    col_name = col_info[1]
                    # 使用PRAGMA中的数据类型，如果为NULL则使用推断类型
                    data_type = col_info[2] if col_info[2] else inferred_types.get(col_name, "unknown")
                    safe_col_id = self._make_safe_id(col_name)

                    col_asset = Column(
                        id=f"sqlite.{safe_db_id}.table.{safe_table_id}.{safe_col_id}",
                        name=col_name,
                        type="column",
                        data_type=data_type,
                        description=f"列: {col_name} in {table_name}",
                        owner="文件采集器",
                        tags=["column", "数据列"],
                        created_time=timestamp,
                        updated_time=timestamp
                    )
                    assets.append(col_asset)

                logging.info(f"表 {table_name} 有 {len(columns_info)} 个列")

                # 6. 行级资产收集
                row_assets = self._collect_sqlite_row_metadata(conn, table_name, safe_db_id, safe_table_id,
                                                               column_names,
                                                               timestamp)
                assets.extend(row_assets)

            conn.close()
            logging.info(f"成功处理SQLite数据库 {file_path.name}，共生成 {len(assets)} 个资产")

        except Exception as e:
            logging.error(f"SQLite数据库处理失败 {file_path}: {e}")
            import traceback
            traceback.print_exc()

        return assets

    def _is_valid_sqlite_file(self, file_path: Path) -> bool:
        """验证文件是否为有效的SQLite数据库"""
        try:
            if not file_path.exists() or file_path.stat().st_size == 0:
                return False

            # 检查SQLite文件头
            with open(file_path, 'rb') as f:
                header = f.read(16)
                # SQLite文件头以"SQLite format 3"开头
                return header.startswith(b'SQLite format 3\000')
        except:
            return False

    def _infer_column_types(self, sample_data: List, column_names: List[str]) -> Dict[str, str]:
        """根据样本数据推断列的数据类型"""
        inferred_types = {}

        if not sample_data:
            return {col: "unknown" for col in column_names}

        for i, col_name in enumerate(column_names):
            values = [row[i] for row in sample_data if i < len(row)]

            if not values:
                inferred_types[col_name] = "unknown"
                continue

            # 检查是否为整数
            if all(isinstance(v, int) and not isinstance(v, bool) for v in values if v is not None):
                inferred_types[col_name] = "INTEGER"
            # 检查是否为浮点数
            elif any(isinstance(v, float) for v in values if v is not None):
                inferred_types[col_name] = "REAL"
            # 检查是否为文本
            elif any(isinstance(v, str) for v in values if v is not None):
                inferred_types[col_name] = "TEXT"
            # 检查是否为布尔值
            elif any(isinstance(v, bool) for v in values if v is not None):
                inferred_types[col_name] = "BOOLEAN"
            # 检查是否为日期时间
            elif any('date' in col_name.lower() or 'time' in col_name.lower() for col_name in column_names):
                inferred_types[col_name] = "DATETIME"
            else:
                inferred_types[col_name] = "TEXT"

        return inferred_types

    def _collect_csv_row_metadata(self, file_path: Path, file_id: str, columns: List[str], timestamp: str) -> List[
        DataRow]:
        """收集CSV行级元数据"""
        row_assets = []
        try:
            enc = self.detect_encoding(file_path)
            df = pd.read_csv(file_path, encoding=enc, nrows=self.sample_rows)

            for idx, row in df.iterrows():
                row_data = row.to_dict()
                row_hash = self._calculate_row_hash(row_data)

                row_asset = DataRow(
                    id=f"file.{file_id}.row_{idx}_{row_hash[:8]}",
                    name=f"行_{idx}",
                    type="row",
                    description=f"数据行 {idx} in {file_path.name}",
                    owner="文件采集器",
                    tags=["data_row", "行数据"],
                    created_time=timestamp,
                    updated_time=timestamp,
                    table_id=f"file.{file_id}",
                    row_hash=row_hash,
                    row_data=row_data,
                    row_index=idx
                )
                row_assets.append(row_asset)

        except Exception as e:
            logging.warning(f"CSV行级元数据收集失败 {file_path}: {e}")

        return row_assets

    def _collect_excel_row_metadata(self, sheet, file_id: str, sheet_id: str, columns: List[str], timestamp: str) -> \
            List[DataRow]:
        """收集Excel行级元数据"""
        row_assets = []
        try:
            # 从第二行开始（跳过标题行），采样指定行数
            max_rows = min(self.sample_rows + 1, sheet.max_row + 1)  # +1 因为从第二行开始

            for row_idx in range(2, max_rows):
                row_data = {}
                for col_idx, col_name in enumerate(columns, 1):
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    # 处理Excel中的特殊类型（如日期）
                    if hasattr(cell_value, 'strftime'):
                        row_data[col_name] = cell_value.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        row_data[col_name] = str(cell_value) if cell_value is not None else ""

                row_hash = self._calculate_row_hash(row_data)
                actual_row_index = row_idx - 2  # 从0开始计数

                row_asset = DataRow(
                    id=f"excel.{file_id}.sheet.{sheet_id}.row_{actual_row_index}_{row_hash[:8]}",
                    name=f"行_{actual_row_index}",
                    type="row",
                    description=f"数据行 {actual_row_index} in {sheet.title}",
                    owner="文件采集器",
                    tags=["data_row", "行数据"],
                    created_time=timestamp,
                    updated_time=timestamp,
                    table_id=f"excel.{file_id}.sheet.{sheet_id}",
                    row_hash=row_hash,
                    row_data=row_data,
                    row_index=actual_row_index
                )
                row_assets.append(row_asset)

        except Exception as e:
            logging.warning(f"Excel行级元数据收集失败 {sheet.title}: {e}")

        return row_assets

    def _collect_sqlite_row_metadata(self, conn, table_name: str, db_id: str, table_id: str, columns: List[str],
                                     timestamp: str) -> List[DataRow]:
        """收集SQLite行级元数据"""
        row_assets = []
        try:
            cursor = conn.cursor()
            cursor.execute(f"SELECT * FROM {table_name} LIMIT {self.sample_rows}")
            rows = cursor.fetchall()

            for idx, row in enumerate(rows):
                row_data = dict(zip(columns, row))
                # 处理SQLite中的特殊类型
                for key, value in row_data.items():
                    if value is None:
                        row_data[key] = ""
                    elif isinstance(value, (int, float)):
                        row_data[key] = value
                    elif isinstance(value, bytes):
                        # 处理BLOB类型数据
                        try:
                            row_data[key] = value.decode('utf-8')
                        except:
                            row_data[key] = f"BLOB_data_{hashlib.md5(value).hexdigest()[:8]}"
                    else:
                        row_data[key] = str(value)

                row_hash = self._calculate_row_hash(row_data)

                row_asset = DataRow(
                    id=f"sqlite.{db_id}.table.{table_id}.row_{idx}_{row_hash[:8]}",
                    name=f"行_{idx}",
                    type="row",
                    description=f"数据行 {idx} in {table_name}",
                    owner="文件采集器",
                    tags=["data_row", "行数据"],
                    created_time=timestamp,
                    updated_time=timestamp,
                    table_id=f"sqlite.{db_id}.table.{table_id}",
                    row_hash=row_hash,
                    row_data=row_data,
                    row_index=idx
                )
                row_assets.append(row_asset)

        except Exception as e:
            logging.warning(f"SQLite行级元数据收集失败 {table_name}: {e}")

        return row_assets

    def detect_encoding(self, file_path: Path) -> str:
        """检测文件编码"""
        for enc in ('utf-8-sig', 'gbk', 'latin1'):
            try:
                with file_path.open('r', encoding=enc) as f:
                    f.read(1024)
                return enc
            except UnicodeDecodeError:
                continue
        return 'utf-8'

    def get_csv_columns(self, file_path: Path) -> List[str]:
        """安全读取 CSV 头"""
        enc = self.detect_encoding(file_path)
        try:
            df = pd.read_csv(file_path, encoding=enc, nrows=0)
            return list(df.columns)
        except Exception as e:
            logging.warning(f"无法读取列信息 {file_path}: {e}")
            return []